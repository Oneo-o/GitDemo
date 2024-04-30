import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class HomePageData:
    sheet: Worksheet
    cell: Cell
    test_HomePage_data = [{"email": "example@example.ex", "password": "Zz123456"},
                          {"email": "example2@example.ex", "password": "Zz123456"}]

    @staticmethod
    def get_test_data(test_case_name):
        my_dict = {}
        book = openpyxl.load_workbook(r"C:\Users\EF\Desktop\Python selenium testing\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    my_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [my_dict]
