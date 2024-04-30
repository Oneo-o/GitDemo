import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

book = openpyxl.load_workbook(r"C:\Users\EF\Desktop\Python selenium testing\PythonDemo.xlsx")
sheet: Worksheet
cell: Cell
sheet = book.active
my_dict = {}
# Read value
cell = sheet.cell(row=1, column=2)
print(cell.value)

# Write value
sheet.cell(row=2, column=2).value = "Vano"

print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)

# Short way to get value of cell
print(sheet['A5'].value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "TestCase2":
        for j in range(2, sheet.max_column + 1):
            my_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(my_dict)
